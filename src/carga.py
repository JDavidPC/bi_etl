"""Rutinas de carga para exportar resultados del pipeline de Airbnb CDMX."""

from __future__ import annotations

import math
import sqlite3
from pathlib import Path
from typing import Optional

import pandas as pd

from logs_manage import Logs


class Carga:
    """Gestiona la persistencia de los datos transformados en diferentes formatos."""

    MAX_EXCEL_ROWS = 1_048_576

    def __init__(
        self,
        df_final,
        df_limpio,
        df_reviews,
        df_calendar_agregado,
        logger: Optional[Logs] = None,
        output_dir: Optional[str | Path] = None,
    ) -> None:
        """Configura la carga con los DataFrames y rutas de salida necesarias."""
        self.df_final = df_final
        self.df_limpio = df_limpio
        self.df_reviews = df_reviews
        self.df_calendar_agregado = df_calendar_agregado
        self.logger = (
            logger.crear_sublogger(header="LOG DE CARGA - AIRBNB CDMX")
            if logger
            else Logs(header="LOG DE CARGA - AIRBNB CDMX")
        )

        default_output = Path(__file__).resolve().parent / "output"
        resolved_dir = Path(output_dir) if output_dir else default_output
        if not resolved_dir.is_absolute():
            resolved_dir = default_output.parent / resolved_dir
        resolved_dir.mkdir(parents=True, exist_ok=True)

        self.output_dir = resolved_dir
        self.sqlite_path = self.output_dir / "bi_mx.db"
        self.xlsx_path = self.output_dir / "datos_limpios.xlsx"

        self.logger.info("Clase Carga inicializada con los DataFrames a exportar")
        self.logger.info(
            f"Formas - final: {df_final.shape}, limpio: {df_limpio.shape}, reviews: {df_reviews.shape}, calendar_agregado: {df_calendar_agregado.shape}"
        )
        self.logger.info(f"Archivos de salida -> {self.output_dir}")

        self._excel_targets = [
            ("listings_limpio", self.df_limpio),
            ("reviews_analizados", self.df_reviews),
            ("calendar_agregado", self.df_calendar_agregado),
        ]
        self._expected_rows = {name: len(frame) for name, frame in self._excel_targets}
        self._excel_registry: dict[str, list[tuple[str, int]]] = {}
        self._sheet_names_in_use: set[str] = set()

    # -------------------------------------------------------------------------
    # Utilidades internas
    # -------------------------------------------------------------------------
    def _resolve_output_path(
        self, file_path: Optional[str | Path], default_path: Path
    ) -> Path:
        """Devuelve una ruta absoluta dentro del directorio de salida."""
        if file_path is None:
            target_path = default_path
        else:
            candidate = Path(file_path)
            target_path = candidate if candidate.is_absolute() else self.output_dir / candidate
        target_path.parent.mkdir(parents=True, exist_ok=True)
        return target_path

    def _unique_sheet_name(self, name: str) -> str:
        """Genera un nombre de hoja válido (<=31 caracteres) y no repetido."""
        sanitized = "".join(ch if ch.isalnum() or ch in (" ", "_") else "_" for ch in name).strip()
        if not sanitized:
            sanitized = "Sheet"
        sanitized = sanitized[:31]

        candidate = sanitized
        counter = 1
        while candidate in self._sheet_names_in_use:
            suffix = f"_{counter}"
            candidate = f"{sanitized[:31 - len(suffix)]}{suffix}"
            counter += 1
        self._sheet_names_in_use.add(candidate)
        return candidate

    def _write_dataframe_to_excel(
        self, writer: pd.ExcelWriter, df: pd.DataFrame, base_name: str
    ) -> int:
        """Exporta un DataFrame creando varias hojas si rebasa el límite de Excel."""
        self._excel_registry[base_name] = []

        if df.empty:
            sheet_name = self._unique_sheet_name(f"{base_name}_sin_datos")
            df.head(0).to_excel(writer, sheet_name=sheet_name, index=False)
            self._excel_registry[base_name].append((sheet_name, 0))
            return 1

        total_rows = len(df)
        if total_rows <= self.MAX_EXCEL_ROWS:
            sheet_name = self._unique_sheet_name(base_name)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            self._excel_registry[base_name].append((sheet_name, total_rows))
            return 1

        segmentos = math.ceil(total_rows / self.MAX_EXCEL_ROWS)
        self.logger.warning(
            f"⚠️  '{base_name}' excede {self.MAX_EXCEL_ROWS:,} filas; se dividirá en {segmentos} hojas."
        )
        for idx in range(segmentos):
            inicio = idx * self.MAX_EXCEL_ROWS
            fin = min((idx + 1) * self.MAX_EXCEL_ROWS, total_rows)
            segmento_df = df.iloc[inicio:fin]
            sheet_name = self._unique_sheet_name(f"{base_name}_{idx + 1}")
            segmento_df.to_excel(writer, sheet_name=sheet_name, index=False)
            self._excel_registry[base_name].append((sheet_name, len(segmento_df)))
        return segmentos

    # -------------------------------------------------------------------------
    # Exportaciones
    # -------------------------------------------------------------------------
    def cargar_sqlite(self, db_path: Optional[str | Path] = None) -> bool:
        """Persiste el DataFrame final en SQLite dentro de la carpeta de salida."""
        target_path = self._resolve_output_path(db_path, self.sqlite_path)
        self.logger.nueva_entrada(f"Carga a SQLite -> {target_path}")
        try:
            with sqlite3.connect(target_path) as conn:
                self.df_final.to_sql("listings_analitica", conn, if_exists="replace", index=False)
            self.logger.info(
                f"Éxito: {len(self.df_final)} registros insertados en 'listings_analitica'"
            )
            return True
        except Exception as exc:  # pragma: no cover - logging de fallos
            self.logger.error(f"Error al cargar datos en SQLite: {exc}")
            return False

    def cargar_xlsx(self, xlsx_path: Optional[str | Path] = None) -> bool:
        """Genera un archivo XLSX en `output` dividiendo hojas cuando es necesario."""
        target_path = self._resolve_output_path(xlsx_path, self.xlsx_path)
        self.logger.nueva_entrada(f"Carga a XLSX -> {target_path}")
        try:
            with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
                for hoja, dataframe in self._excel_targets:
                    segmentos = self._write_dataframe_to_excel(writer, dataframe, hoja)
                    if segmentos > 1:
                        self.logger.info(
                            f"'{hoja}' se exportó en {segmentos} hojas por límite de Excel."
                        )
            total_hojas = sum(len(info) for info in self._excel_registry.values())
            self.logger.info(
                f"Éxito: Se guardaron {total_hojas} hojas en el archivo '{target_path.name}'"
            )
            return True
        except Exception as exc:  # pragma: no cover - logging de fallos
            self.logger.error(f"Error al guardar datos en XLSX: {exc}")
            return False

    # -------------------------------------------------------------------------
    # Verificaciones
    # -------------------------------------------------------------------------
    def verificar_carga(
        self,
        db_path: Optional[str | Path] = None,
        xlsx_path: Optional[str | Path] = None,
    ) -> None:
        """Comprueba que los archivos generados tengan la cantidad esperada de filas."""
        self.logger.nueva_entrada("Verificación de cargas")

        try:
            target_path = self._resolve_output_path(db_path, self.sqlite_path)
            with sqlite3.connect(target_path) as conn:
                count_sqlite = pd.read_sql_query(
                    "SELECT COUNT(*) FROM listings_analitica", conn
                ).iloc[0, 0]
            if count_sqlite == len(self.df_final):
                self.logger.info(f"✅ SQLite ok ({count_sqlite} registros)")
            else:
                self.logger.warning(
                    f"❌ SQLite mismatch (Esperado: {len(self.df_final)}, Encontrado: {count_sqlite})"
                )
        except Exception as exc:  # pragma: no cover - logging de fallos
            self.logger.error(f"❌ Error al verificar SQLite: {exc}")

        try:
            target_path = self._resolve_output_path(xlsx_path, self.xlsx_path)
            if not target_path.exists():
                raise FileNotFoundError(target_path)

            try:
                from openpyxl import load_workbook
            except ImportError as exc:  # pragma: no cover - depende de entorno
                self.logger.warning(f"No se pudo verificar XLSX: falta openpyxl ({exc})")
                return

            workbook = load_workbook(target_path, read_only=True)
            try:
                for base_name, sheets in self._excel_registry.items():
                    if not sheets:
                        continue

                    total_rows = 0
                    for sheet_name, expected_rows in sheets:
                        if sheet_name not in workbook.sheetnames:
                            self.logger.warning(
                                f"❌ Hoja '{sheet_name}' no encontrada para '{base_name}'"
                            )
                            break
                        hoja = workbook[sheet_name]
                        actual_rows = max(hoja.max_row - 1, 0)
                        if actual_rows != expected_rows:
                            self.logger.warning(
                                f"❌ XLSX mismatch en '{sheet_name}' (Esperado: {expected_rows}, Encontrado: {actual_rows})"
                            )
                            break
                        total_rows += actual_rows
                    else:
                        esperado = self._expected_rows.get(base_name, 0)
                        if total_rows == esperado:
                            self.logger.info(
                                f"✅ XLSX '{base_name}' ok ({total_rows} registros en {len(sheets)} hojas)"
                            )
                        else:
                            self.logger.warning(
                                f"❌ XLSX mismatch en '{base_name}' (Esperado total: {esperado}, Encontrado: {total_rows})"
                            )
            finally:
                workbook.close()
        except Exception as exc:  # pragma: no cover - logging de fallos
            self.logger.error(f"❌ Error al verificar XLSX: {exc}")

        self.logger.info("Verificación finalizada")

    # -------------------------------------------------------------------------
    # Operación completa
    # -------------------------------------------------------------------------
    def ejecutar_carga_completa(self) -> None:
        """Ejecuta la carga en todos los destinos y realiza las verificaciones."""
        self.logger.nueva_entrada("Inicio de carga completa")
        sqlite_ok = self.cargar_sqlite()
        xlsx_ok = self.cargar_xlsx()

        if sqlite_ok or xlsx_ok:
            self.verificar_carga()
        else:
            self.logger.warning("No se realizó ninguna carga exitosa. Verificación omitida.")
        self.logger.info("Proceso de carga finalizado")
